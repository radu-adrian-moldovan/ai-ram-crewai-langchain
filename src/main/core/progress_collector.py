#!/usr/bin/env python3
"""
Progress Collector System with Event-Driven Architecture and Excel Reporting
Implements an MCP-like event system for tracking document processing progress
"""

import asyncio
import time
import uuid
from datetime import datetime, timezone
from enum import Enum
from typing import Dict, List, Optional, Any, Union
from dataclasses import dataclass, field
from pathlib import Path
import json

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference


class ImplementationType(Enum):
    """Implementation types for tracking"""
    CREWAI = "crewai"
    LANGCHAIN = "langchain"


class EventType(Enum):
    """Event types for progress tracking"""
    PROCESSING_START = "processing_start"
    PROCESSING_END = "processing_end"
    CLASSIFICATION_START = "classification_start"
    CLASSIFICATION_END = "classification_end"
    EXTRACTION_START = "extraction_start"
    EXTRACTION_END = "extraction_end"
    ACCURACY_ASSESSMENT_START = "accuracy_assessment_start"
    ACCURACY_ASSESSMENT_END = "accuracy_assessment_end"
    FINAL_DECISION_START = "final_decision_start"
    FINAL_DECISION_END = "final_decision_end"
    ERROR = "error"
    WARNING = "warning"
    INFO = "info"


@dataclass
class ProcessingEvent:
    """Event data structure for tracking processing events"""
    event_id: str = field(default_factory=lambda: str(uuid.uuid4()))
    timestamp: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    implementation: Optional[ImplementationType] = None
    event_type: Optional[EventType] = None
    document_path: str = ""
    document_name: str = ""
    stage: str = ""
    duration_ms: Optional[float] = None
    success: bool = True
    error_message: Optional[str] = None
    
    # Processing results
    classification_result: Optional[Dict[str, Any]] = None
    extraction_result: Optional[Dict[str, Any]] = None
    accuracy_result: Optional[Dict[str, Any]] = None
    final_decision_result: Optional[Dict[str, Any]] = None
    
    # Performance metrics
    memory_usage_mb: Optional[float] = None
    cpu_usage_percent: Optional[float] = None
    
    # Additional metadata
    metadata: Dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> Dict[str, Any]:
        """Convert event to dictionary for serialization"""
        return {
            'event_id': self.event_id,
            'timestamp': self.timestamp.isoformat(),
            'implementation': self.implementation.value if self.implementation else None,
            'event_type': self.event_type.value if self.event_type else None,
            'document_path': self.document_path,
            'document_name': self.document_name,
            'stage': self.stage,
            'duration_ms': self.duration_ms,
            'success': self.success,
            'error_message': self.error_message,
            'classification_result': self.classification_result,
            'extraction_result': self.extraction_result,
            'accuracy_result': self.accuracy_result,
            'final_decision_result': self.final_decision_result,
            'memory_usage_mb': self.memory_usage_mb,
            'cpu_usage_percent': self.cpu_usage_percent,
            'metadata': self.metadata
        }


class EventCollector:
    """Event collector for managing processing events"""
    
    def __init__(self):
        self.events: Dict[ImplementationType, List[ProcessingEvent]] = {
            ImplementationType.CREWAI: [],
            ImplementationType.LANGCHAIN: []
        }
        self.session_id = str(uuid.uuid4())
        self.session_start_time = datetime.now(timezone.utc)
        
    def add_event(self, event: ProcessingEvent):
        """Add an event to the appropriate topic"""
        if event.implementation:
            self.events[event.implementation].append(event)
    
    def get_events(self, implementation: Optional[ImplementationType] = None) -> List[ProcessingEvent]:
        """Get events for a specific implementation or all events"""
        if implementation:
            return self.events[implementation].copy()
        
        all_events = []
        for impl_events in self.events.values():
            all_events.extend(impl_events)
        return sorted(all_events, key=lambda e: e.timestamp)
    
    def get_session_summary(self) -> Dict[str, Any]:
        """Get session summary statistics"""
        total_events = sum(len(events) for events in self.events.values())
        session_duration = (datetime.now(timezone.utc) - self.session_start_time).total_seconds()
        
        return {
            'session_id': self.session_id,
            'start_time': self.session_start_time,
            'total_events': total_events,
            'session_duration_seconds': session_duration,
            'events_by_implementation': {
                impl.value: len(events) for impl, events in self.events.items()
            }
        }


class PerformanceMetricsCalculator:
    """Calculate performance metrics from events"""
    
    @staticmethod
    def calculate_processing_metrics(events: List[ProcessingEvent]) -> Dict[str, Any]:
        """Calculate processing performance metrics"""
        if not events:
            return {}
        
        # Filter processing events
        start_events = [e for e in events if e.event_type == EventType.PROCESSING_START]
        end_events = [e for e in events if e.event_type == EventType.PROCESSING_END]
        error_events = [e for e in events if e.event_type == EventType.ERROR]
        
        # Calculate timing metrics
        processing_times = []
        successful_docs = 0
        failed_docs = len(error_events)
        
        # Match start/end events by document
        for start_event in start_events:
            matching_end = next(
                (e for e in end_events 
                 if e.document_path == start_event.document_path and e.timestamp > start_event.timestamp),
                None
            )
            
            if matching_end:
                duration = (matching_end.timestamp - start_event.timestamp).total_seconds()
                processing_times.append(duration)
                if matching_end.success:
                    successful_docs += 1
        
        total_docs = len(start_events)
        success_rate = successful_docs / total_docs if total_docs > 0 else 0
        
        return {
            'total_documents': total_docs,
            'successful_documents': successful_docs,
            'failed_documents': failed_docs,
            'success_rate': success_rate,
            'total_processing_time': sum(processing_times),
            'average_processing_time': sum(processing_times) / len(processing_times) if processing_times else 0,
            'min_processing_time': min(processing_times) if processing_times else 0,
            'max_processing_time': max(processing_times) if processing_times else 0,
            'processing_times': processing_times
        }
    
    @staticmethod
    def calculate_accuracy_metrics(events: List[ProcessingEvent]) -> Dict[str, Any]:
        """Calculate accuracy metrics from events"""
        accuracy_events = [e for e in events if e.accuracy_result]
        
        if not accuracy_events:
            return {}
        
        accuracies = [e.accuracy_result.get('overall_accuracy', 0) for e in accuracy_events]
        confidences = []
        
        # Extract confidence scores from classification results
        for event in events:
            if event.classification_result and 'confidence' in event.classification_result:
                confidences.append(event.classification_result['confidence'])
        
        return {
            'total_assessments': len(accuracy_events),
            'average_accuracy': sum(accuracies) / len(accuracies) if accuracies else 0,
            'min_accuracy': min(accuracies) if accuracies else 0,
            'max_accuracy': max(accuracies) if accuracies else 0,
            'average_confidence': sum(confidences) / len(confidences) if confidences else 0,
            'accuracy_scores': accuracies,
            'confidence_scores': confidences
        }


class ProgressCollectorMCP:
    """Main Progress Collector class implementing MCP-like functionality"""
    
    def __init__(self):
        self.collector = EventCollector()
        self.active_timers: Dict[str, datetime] = {}
    
    def start_processing(self, implementation: ImplementationType, document_path: str, metadata: Dict[str, Any] = None) -> str:
        """Start processing timer and create start event"""
        document_name = Path(document_path).name
        timer_key = f"{implementation.value}:{document_path}"
        
        event = ProcessingEvent(
            implementation=implementation,
            event_type=EventType.PROCESSING_START,
            document_path=document_path,
            document_name=document_name,
            stage="processing",
            metadata=metadata or {}
        )
        
        self.collector.add_event(event)
        self.active_timers[timer_key] = event.timestamp
        
        return event.event_id
    
    def end_processing(self, implementation: ImplementationType, document_path: str, 
                      success: bool = True, error_message: str = None, result: Any = None) -> str:
        """End processing timer and create end event"""
        document_name = Path(document_path).name
        timer_key = f"{implementation.value}:{document_path}"
        
        # Calculate duration
        duration_ms = None
        if timer_key in self.active_timers:
            start_time = self.active_timers[timer_key]
            duration_ms = (datetime.now(timezone.utc) - start_time).total_seconds() * 1000
            del self.active_timers[timer_key]
        
        # Extract results if available
        classification_result = None
        extraction_result = None
        accuracy_result = None
        final_decision_result = None
        
        # Check if result is a dictionary (our enhanced format)
        if isinstance(result, dict):
            # Extract enhanced JSON data
            classification_result = result.get('classification_json')
            extraction_result = result.get('extraction_json')
            accuracy_result = result.get('accuracy_json')
            final_decision_result = result.get('final_decision_json')
        else:
            # Legacy format - extract from object attributes
            if result and hasattr(result, 'classification'):
                classification_result = {
                    'document_type': result.classification.document_type,
                    'confidence': result.classification.confidence,
                    'reasoning': getattr(result.classification, 'reasoning', '')
                }
            
            if result and hasattr(result, 'extraction'):
                extraction_result = {
                    'total_fields_found': result.extraction.total_fields_found,
                    'extracted_fields': [
                        {
                            'field_name': field.field_name,
                            'value': field.value,
                            'confidence': getattr(field, 'confidence', 0.0)
                        } for field in result.extraction.extracted_fields[:5]  # Limit for performance
                    ]
                }
            
            if result and hasattr(result, 'accuracy_assessment'):
                accuracy_result = {
                    'overall_accuracy': result.accuracy_assessment.overall_accuracy,
                    'issues_found': getattr(result.accuracy_assessment, 'issues_found', [])
                }
            
            if result and hasattr(result, 'final_decision'):
                final_decision_result = {
                    'accept': result.final_decision.accept,
                    'confidence': result.final_decision.confidence,
                    'reasoning': getattr(result.final_decision, 'reasoning', '')
                }
        
        event = ProcessingEvent(
            implementation=implementation,
            event_type=EventType.PROCESSING_END,
            document_path=document_path,
            document_name=document_name,
            stage="processing",
            duration_ms=duration_ms,
            success=success,
            error_message=error_message,
            classification_result=classification_result,
            extraction_result=extraction_result,
            accuracy_result=accuracy_result,
            final_decision_result=final_decision_result
        )
        
        self.collector.add_event(event)
        return event.event_id
    
    def log_stage_event(self, implementation: ImplementationType, document_path: str, 
                       stage: str, event_type: EventType, success: bool = True, 
                       error_message: str = None, metadata: Dict[str, Any] = None) -> str:
        """Log a stage-specific event"""
        document_name = Path(document_path).name
        
        # Extract stage-specific results from metadata and map to proper fields
        classification_result = None
        extraction_result = None
        accuracy_result = None
        
        if event_type == EventType.CLASSIFICATION_END and metadata:
            # Move classification data from metadata to classification_result
            classification_result = {
                'document_type': metadata.get('document_type'),
                'confidence': metadata.get('confidence'),
                'reasoning': metadata.get('reasoning', f"Classified as {metadata.get('document_type', 'unknown')}")
            }
            
        elif event_type == EventType.EXTRACTION_END and metadata:
            # Extract field data for extraction_result
            extracted_fields = []
            for k, v in metadata.items():
                if k not in ['confidence', 'processing_time']:
                    extracted_fields.append({
                        'field_name': k, 
                        'value': str(v), 
                        'confidence': metadata.get('confidence', 0.8)
                    })
            
            extraction_result = {
                'extracted_fields': extracted_fields,
                'total_fields_found': len(extracted_fields)
            }
            
        elif event_type == EventType.ACCURACY_ASSESSMENT_END and metadata:
            accuracy_result = {
                'overall_accuracy': metadata.get('overall_accuracy', 0.0),
                'field_accuracies': metadata.get('field_accuracies', {}),
                'issues_found': metadata.get('issues_found', []),
                'recommendations': metadata.get('recommendations', [])
            }
        
        event = ProcessingEvent(
            implementation=implementation,
            event_type=event_type,
            document_path=document_path,
            document_name=document_name,
            stage=stage,
            success=success,
            error_message=error_message,
            classification_result=classification_result,
            extraction_result=extraction_result,
            accuracy_result=accuracy_result,
            metadata=metadata or {}
        )
        
        self.collector.add_event(event)
        return event.event_id
    
    def log_error(self, implementation: ImplementationType, document_path: str, 
                  error_message: str, stage: str = "unknown") -> str:
        """Log an error event"""
        return self.log_stage_event(
            implementation=implementation,
            document_path=document_path,
            stage=stage,
            event_type=EventType.ERROR,
            success=False,
            error_message=error_message
        )
    
    def log_info(self, implementation: ImplementationType, document_path: str, 
                 message: str, stage: str = "info") -> str:
        """Log an info event"""
        return self.log_stage_event(
            implementation=implementation,
            document_path=document_path,
            stage=stage,
            event_type=EventType.INFO,
            metadata={'message': message}
        )
    
    def get_summary_statistics(self) -> Dict[str, Any]:
        """Get comprehensive summary statistics"""
        session_info = self.collector.get_session_summary()
        
        crewai_events = self.collector.get_events(ImplementationType.CREWAI)
        langchain_events = self.collector.get_events(ImplementationType.LANGCHAIN)
        
        metrics_calc = PerformanceMetricsCalculator()
        crewai_metrics = metrics_calc.calculate_processing_metrics(crewai_events)
        langchain_metrics = metrics_calc.calculate_processing_metrics(langchain_events)
        
        crewai_accuracy = metrics_calc.calculate_accuracy_metrics(crewai_events)
        langchain_accuracy = metrics_calc.calculate_accuracy_metrics(langchain_events)
        
        return {
            'session_info': session_info,
            'crewai_metrics': crewai_metrics,
            'langchain_metrics': langchain_metrics,
            'crewai_accuracy': crewai_accuracy,
            'langchain_accuracy': langchain_accuracy
        }
    
    def generate_excel_report(self, output_path: str) -> str:
        """Generate comprehensive Excel report"""
        return self._create_excel_report(output_path)
    
    def _create_excel_report(self, output_path: str) -> str:
        """Create Excel report with multiple sheets"""
        workbook = Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Create sheets
        self._create_executive_summary_sheet(workbook)
        self._create_detailed_events_sheet(workbook)
        self._create_performance_metrics_sheet(workbook)
        self._create_enhanced_results_sheet(workbook)
        self._create_trace_analysis_sheet(workbook)
        
        # Save workbook
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(str(output_path))
        
        return str(output_path)
    
    def _create_executive_summary_sheet(self, workbook: Workbook):
        """Create executive summary sheet"""
        ws = workbook.create_sheet("Executive Summary")
        
        # Session information
        session_info = self.collector.get_session_summary()
        
        # Calculate metrics for both implementations
        crewai_events = self.collector.get_events(ImplementationType.CREWAI)
        langchain_events = self.collector.get_events(ImplementationType.LANGCHAIN)
        
        metrics_calc = PerformanceMetricsCalculator()
        crewai_metrics = metrics_calc.calculate_processing_metrics(crewai_events)
        langchain_metrics = metrics_calc.calculate_processing_metrics(langchain_events)
        
        crewai_accuracy = metrics_calc.calculate_accuracy_metrics(crewai_events)
        langchain_accuracy = metrics_calc.calculate_accuracy_metrics(langchain_events)
        
        # Headers and styling
        header_font = Font(size=14, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Title
        ws['A1'] = "Document Processing Comparison Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        # Session Information
        row = 3
        ws[f'A{row}'] = "Session Information"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws[f'A{row}'] = "Session ID:"
        ws[f'B{row}'] = session_info['session_id']
        
        row += 1
        ws[f'A{row}'] = "Start Time:"
        ws[f'B{row}'] = session_info['start_time'].strftime("%Y-%m-%d %H:%M:%S UTC")
        
        row += 1
        ws[f'A{row}'] = "Total Events:"
        ws[f'B{row}'] = session_info['total_events']
        
        # Performance Comparison
        row += 2
        ws[f'A{row}'] = "Performance Comparison"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        headers = ["Metric", "CrewAI", "LangChain", "Winner"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Add metrics rows
        metrics_data = [
            ("Documents Processed", crewai_metrics.get('total_documents', 0), langchain_metrics.get('total_documents', 0)),
            ("Success Rate", f"{crewai_metrics.get('success_rate', 0):.1%}", f"{langchain_metrics.get('success_rate', 0):.1%}"),
            ("Avg Processing Time (s)", f"{crewai_metrics.get('average_processing_time', 0):.2f}", f"{langchain_metrics.get('average_processing_time', 0):.2f}"),
            ("Total Processing Time (s)", f"{crewai_metrics.get('total_processing_time', 0):.2f}", f"{langchain_metrics.get('total_processing_time', 0):.2f}"),
            ("Avg Accuracy", f"{crewai_accuracy.get('average_accuracy', 0):.2f}", f"{langchain_accuracy.get('average_accuracy', 0):.2f}"),
            ("Avg Confidence", f"{crewai_accuracy.get('average_confidence', 0):.2f}", f"{langchain_accuracy.get('average_confidence', 0):.2f}"),
        ]
        
        for metric_name, crewai_val, langchain_val in metrics_data:
            row += 1
            ws[f'A{row}'] = metric_name
            ws[f'B{row}'] = crewai_val
            ws[f'C{row}'] = langchain_val
            
            # Determine winner (simple logic for demo)
            if "Time" in metric_name:
                try:
                    crewai_num = float(str(crewai_val).replace('s', ''))
                    langchain_num = float(str(langchain_val).replace('s', ''))
                    winner = "CrewAI" if crewai_num < langchain_num else "LangChain"
                except:
                    winner = "Tie"
            else:
                try:
                    crewai_num = float(str(crewai_val).replace('%', ''))
                    langchain_num = float(str(langchain_val).replace('%', ''))
                    winner = "CrewAI" if crewai_num > langchain_num else "LangChain"
                except:
                    winner = "Tie"
            ws[f'D{row}'] = winner
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            # Find the first regular cell to get column letter
            for cell in column:
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                    break
            
            if column_letter is None:
                continue
                
            for cell in column:
                try:
                    if hasattr(cell, 'value') and cell.value is not None:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_detailed_events_sheet(self, workbook: Workbook):
        """Create detailed events sheet"""
        ws = workbook.create_sheet("Detailed Events")
        
        # Get all events and convert to DataFrame
        all_events = self.collector.get_events()
        
        if not all_events:
            ws['A1'] = "No events recorded"
            return
        
        # Convert events to rows
        rows_data = []
        for event in all_events:
            row_data = {
                'Timestamp': event.timestamp.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
                'Implementation': event.implementation.value if event.implementation else '',
                'Event Type': event.event_type.value if event.event_type else '',
                'Document Name': event.document_name,
                'Stage': event.stage,
                'Duration (ms)': event.duration_ms if event.duration_ms else '',
                'Success': 'Yes' if event.success else 'No',
                'Error Message': event.error_message if event.error_message else '',
                'Classification Type': event.classification_result.get('classification', {}).get('label', '') if event.classification_result and isinstance(event.classification_result, dict) else '',
                'Classification Confidence': event.classification_result.get('classification', {}).get('confidence', '') if event.classification_result and isinstance(event.classification_result, dict) else '',
                'Classification JSON': json.dumps(event.classification_result, ensure_ascii=False, indent=2, separators=(',', ': ')) if event.classification_result else '',
                'Extraction JSON': json.dumps(event.extraction_result, ensure_ascii=False, indent=2, separators=(',', ': ')) if event.extraction_result else '',
                'Accuracy JSON': json.dumps(event.accuracy_result, ensure_ascii=False, indent=2, separators=(',', ': ')) if event.accuracy_result else '',
                'Final Decision JSON': json.dumps(event.final_decision_result, ensure_ascii=False, indent=2, separators=(',', ': ')) if event.final_decision_result else '',
                'Overall Accuracy': event.accuracy_result.get('overall', '') if event.accuracy_result and isinstance(event.accuracy_result, dict) else '',
                'Total Processing Time': event.final_decision_result.get('timing', {}).get('total_processing_time', '') if event.final_decision_result and isinstance(event.final_decision_result, dict) else '',
                'Framework Processing Time': event.final_decision_result.get('timing', {}).get('framework_processing_time', '') if event.final_decision_result and isinstance(event.final_decision_result, dict) else '',
            }
            rows_data.append(row_data)
        
        # Create DataFrame and write to sheet
        df = pd.DataFrame(rows_data)
        
        # Write headers
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Write data
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                # Format JSON cells with text wrapping and proper alignment
                column_title = df.columns[col_num - 1]
                if 'JSON' in column_title:
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            # Find the first regular cell to get column letter
            for cell in column:
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                    break
            
            if column_letter is None:
                continue
                
            for cell in column:
                try:
                    if hasattr(cell, 'value') and cell.value is not None:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_performance_metrics_sheet(self, workbook: Workbook):
        """Create performance metrics sheet"""
        ws = workbook.create_sheet("Performance Metrics")
        
        # Calculate metrics for both implementations
        crewai_events = self.collector.get_events(ImplementationType.CREWAI)
        langchain_events = self.collector.get_events(ImplementationType.LANGCHAIN)
        
        metrics_calc = PerformanceMetricsCalculator()
        crewai_metrics = metrics_calc.calculate_processing_metrics(crewai_events)
        langchain_metrics = metrics_calc.calculate_processing_metrics(langchain_events)
        
        crewai_accuracy = metrics_calc.calculate_accuracy_metrics(crewai_events)
        langchain_accuracy = metrics_calc.calculate_accuracy_metrics(langchain_events)
        
        # Write headers
        headers = ["Metric", "CrewAI", "LangChain", "Difference", "Better"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Processing metrics
        processing_metrics = [
            ("Total Documents", crewai_metrics.get('total_documents', 0), langchain_metrics.get('total_documents', 0)),
            ("Successful Documents", crewai_metrics.get('successful_documents', 0), langchain_metrics.get('successful_documents', 0)),
            ("Failed Documents", crewai_metrics.get('failed_documents', 0), langchain_metrics.get('failed_documents', 0)),
            ("Success Rate (%)", crewai_metrics.get('success_rate', 0) * 100, langchain_metrics.get('success_rate', 0) * 100),
            ("Total Processing Time (s)", crewai_metrics.get('total_processing_time', 0), langchain_metrics.get('total_processing_time', 0)),
            ("Average Processing Time (s)", crewai_metrics.get('average_processing_time', 0), langchain_metrics.get('average_processing_time', 0)),
            ("Min Processing Time (s)", crewai_metrics.get('min_processing_time', 0), langchain_metrics.get('min_processing_time', 0)),
            ("Max Processing Time (s)", crewai_metrics.get('max_processing_time', 0), langchain_metrics.get('max_processing_time', 0)),
        ]
        
        # Accuracy metrics
        accuracy_metrics = [
            ("Average Accuracy", crewai_accuracy.get('average_accuracy', 0), langchain_accuracy.get('average_accuracy', 0)),
            ("Min Accuracy", crewai_accuracy.get('min_accuracy', 0), langchain_accuracy.get('min_accuracy', 0)),
            ("Max Accuracy", crewai_accuracy.get('max_accuracy', 0), langchain_accuracy.get('max_accuracy', 0)),
            ("Average Confidence", crewai_accuracy.get('average_confidence', 0), langchain_accuracy.get('average_confidence', 0)),
        ]
        
        # Write processing metrics
        row = 2
        ws[f'A{row}'] = "PROCESSING METRICS"
        ws[f'A{row}'].font = Font(bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        
        for metric_name, crewai_val, langchain_val in processing_metrics:
            row += 1
            ws[f'A{row}'] = metric_name
            ws[f'B{row}'] = round(crewai_val, 3) if isinstance(crewai_val, (int, float)) else crewai_val
            ws[f'C{row}'] = round(langchain_val, 3) if isinstance(langchain_val, (int, float)) else langchain_val
            
            if isinstance(crewai_val, (int, float)) and isinstance(langchain_val, (int, float)):
                diff = crewai_val - langchain_val
                ws[f'D{row}'] = round(diff, 3)
                
                # Determine which is better (lower time is better, higher success/accuracy is better)
                if "Time" in metric_name:
                    ws[f'E{row}'] = "CrewAI" if crewai_val < langchain_val else "LangChain"
                else:
                    ws[f'E{row}'] = "CrewAI" if crewai_val > langchain_val else "LangChain"
        
        # Write accuracy metrics
        row += 2
        ws[f'A{row}'] = "ACCURACY METRICS"
        ws[f'A{row}'].font = Font(bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        
        for metric_name, crewai_val, langchain_val in accuracy_metrics:
            row += 1
            ws[f'A{row}'] = metric_name
            ws[f'B{row}'] = round(crewai_val, 3) if isinstance(crewai_val, (int, float)) else crewai_val
            ws[f'C{row}'] = round(langchain_val, 3) if isinstance(langchain_val, (int, float)) else langchain_val
            
            if isinstance(crewai_val, (int, float)) and isinstance(langchain_val, (int, float)):
                diff = crewai_val - langchain_val
                ws[f'D{row}'] = round(diff, 3)
                ws[f'E{row}'] = "CrewAI" if crewai_val > langchain_val else "LangChain"
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            # Find the first regular cell to get column letter
            for cell in column:
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                    break
            
            if column_letter is None:
                continue
                
            for cell in column:
                try:
                    if hasattr(cell, 'value') and cell.value is not None:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_enhanced_results_sheet(self, workbook: Workbook):
        """Create enhanced results sheet with JSON trace information"""
        ws = workbook.create_sheet("Enhanced Results & Traces")
        
        # Headers and styling
        header_font = Font(size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Define columns for enhanced results
        headers = [
            "Document", "Framework", "Status", "Processing Time (ms)", 
            "Overall Accuracy", "Classification", "Classification Confidence",
            "Extraction Fields Count", "Step 1: Classification (ms)", "Step 2: Extraction (ms)",
            "Step 3: Accuracy (ms)", "Step 4: Decision (ms)", "Trace Route Summary",
            "Classification JSON", "Extraction JSON", "Accuracy JSON", "Decision JSON"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        
        # Get enhanced events with trace information
        all_events = self.collector.get_events()
        enhanced_events = [e for e in all_events if e.event_type == EventType.PROCESSING_END 
                          and e.metadata.get('enhanced_result')]
        
        row = 2
        for event in enhanced_events:
            if not event.implementation:
                continue
                
            enhanced_result = event.metadata.get('enhanced_result', {})
            processing_pipeline = enhanced_result.get('processing_pipeline', [])
            
            # Extract step timings
            step_timings = {}
            for step in processing_pipeline:
                step_name = step.get('step', 'unknown')
                duration_ms = step.get('duration_ms', 0)
                step_timings[step_name] = duration_ms
            
            # Write basic information
            ws.cell(row=row, column=1).value = event.document_name
            ws.cell(row=row, column=2).value = event.implementation.value
            ws.cell(row=row, column=3).value = enhanced_result.get('status', 'unknown')
            
            # Processing time
            total_time_ms = enhanced_result.get('total_processing_time', 0) * 1000
            ws.cell(row=row, column=4).value = f"{total_time_ms:.2f}"
            
            # Accuracy and confidence
            ws.cell(row=row, column=5).value = f"{enhanced_result.get('overall_accuracy', 0):.3f}"
            ws.cell(row=row, column=6).value = enhanced_result.get('document_type', 'unknown')
            ws.cell(row=row, column=7).value = f"{enhanced_result.get('confidence', 0):.3f}"
            
            # Extraction fields count
            extraction_json = enhanced_result.get('extraction_json', {})
            fields_count = 0
            if extraction_json and 'fields' in extraction_json:
                fields_count = len(extraction_json['fields'])
            elif extraction_json and 'extraction' in extraction_json and 'fields' in extraction_json['extraction']:
                fields_count = len(extraction_json['extraction']['fields'])
            ws.cell(row=row, column=8).value = fields_count
            
            # Step timings
            ws.cell(row=row, column=9).value = f"{step_timings.get('classification', 0):.2f}"
            ws.cell(row=row, column=10).value = f"{step_timings.get('extraction', 0):.2f}"
            ws.cell(row=row, column=11).value = f"{step_timings.get('accuracy_assessment', 0):.2f}"
            ws.cell(row=row, column=12).value = f"{step_timings.get('final_decision', 0):.2f}"
            
            # Trace route summary
            trace_summary = f"Total steps: {len(processing_pipeline)}, "
            successful_steps = len([s for s in processing_pipeline if s.get('status') == 'completed'])
            trace_summary += f"Successful: {successful_steps}, "
            trace_summary += f"Total time: {total_time_ms:.0f}ms"
            ws.cell(row=row, column=13).value = trace_summary
            
            # JSON outputs (full JSON with enhanced formatting for Excel readability)
            def format_json_for_excel(json_obj):
                if not json_obj:
                    return "N/A"
                # Pretty print JSON with enhanced formatting for Excel cells
                # Use smaller indent for better cell display and ensure line breaks
                formatted_json = json.dumps(json_obj, ensure_ascii=False, indent=2, separators=(',', ': '))
                # For Excel, we need to ensure the JSON is properly formatted with line breaks
                # Excel handles actual newlines better than escaped \n characters
                return formatted_json
            
            # Apply JSON formatting and set specific JSON cells
            json_cells = [
                (row, 14, enhanced_result.get('classification_json')),
                (row, 15, enhanced_result.get('extraction_json')),
                (row, 16, enhanced_result.get('accuracy_json')),
                (row, 17, enhanced_result.get('final_decision_json'))
            ]
            
            for cell_row, cell_col, json_data in json_cells:
                cell = ws.cell(row=cell_row, column=cell_col)
                formatted_json = format_json_for_excel(json_data)
                cell.value = formatted_json
                
                # Enhanced formatting for JSON cells
                cell.alignment = Alignment(
                    wrap_text=True, 
                    vertical='top', 
                    horizontal='left'
                )
                # Set font to monospace for better JSON readability with slightly larger size
                cell.font = Font(name='Courier New', size=10)
                
                # Dynamically adjust row height based on JSON content
                if formatted_json and formatted_json != "N/A":
                    line_count = formatted_json.count('\n') + 1
                    # Set row height based on JSON lines (minimum 30, scale with content)
                    target_height = max(30, min(line_count * 14, 400))
                    current_height = ws.row_dimensions[cell_row].height or 15
                    ws.row_dimensions[cell_row].height = max(current_height, target_height)
            
            row += 1
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
        # Add borders and formatting
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row_cells in ws.iter_rows(min_row=1, max_row=row-1, min_col=1, max_col=len(headers)):
            for cell in row_cells:
                cell.border = thin_border
                if cell.row > 1:  # Data rows
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    def _create_trace_analysis_sheet(self, workbook: Workbook):
        """Create trace analysis sheet with performance breakdown"""
        ws = workbook.create_sheet("Trace Analysis")
        
        # Headers and styling
        header_font = Font(size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Get all enhanced events
        all_events = self.collector.get_events()
        enhanced_events = [e for e in all_events if e.event_type == EventType.PROCESSING_END 
                          and e.metadata.get('enhanced_result')]
        
        if not enhanced_events:
            ws['A1'] = "No enhanced trace data available"
            return
        
        # Analyze step performance across frameworks
        step_analysis = {}
        framework_totals = {}
        
        for event in enhanced_events:
            if not event.implementation:
                continue
                
            framework = event.implementation.value
            enhanced_result = event.metadata.get('enhanced_result', {})
            processing_pipeline = enhanced_result.get('processing_pipeline', [])
            
            if framework not in framework_totals:
                framework_totals[framework] = {
                    'total_time_ms': 0,
                    'document_count': 0,
                    'step_times': {}
                }
            
            total_doc_time = 0
            for step in processing_pipeline:
                step_name = step.get('step', 'unknown')
                duration_ms = step.get('duration_ms', 0)
                total_doc_time += duration_ms
                
                if step_name not in step_analysis:
                    step_analysis[step_name] = {}
                if framework not in step_analysis[step_name]:
                    step_analysis[step_name][framework] = []
                
                step_analysis[step_name][framework].append(duration_ms)
                
                if step_name not in framework_totals[framework]['step_times']:
                    framework_totals[framework]['step_times'][step_name] = []
                framework_totals[framework]['step_times'][step_name].append(duration_ms)
            
            framework_totals[framework]['total_time_ms'] += total_doc_time
            framework_totals[framework]['document_count'] += 1
        row = 1
        # Create step performance comparison
        row = 1
        ws[f'A{row}'] = "Step Performance Analysis"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 2
        headers = ["Step", "CrewAI Avg (ms)", "LangChain Avg (ms)", "CrewAI Min (ms)", 
                  "CrewAI Max (ms)", "Performance Winner"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        
        row += 1
        for step_name, framework_data in step_analysis.items():
            ws.cell(row=row, column=1).value = step_name.replace('_', ' ').title()
            
            crewai_times = framework_data.get('crewai', [])
            langchain_times = framework_data.get('langchain', [])
            
            crewai_avg = sum(crewai_times) / len(crewai_times) if crewai_times else 0
            langchain_avg = sum(langchain_times) / len(langchain_times) if langchain_times else 0
            
            ws.cell(row=row, column=2).value = f"{crewai_avg:.2f}"
            ws.cell(row=row, column=3).value = f"{langchain_avg:.2f}"
            
            if crewai_times:
                ws.cell(row=row, column=4).value = f"{min(crewai_times):.2f}"
                ws.cell(row=row, column=5).value = f"{max(crewai_times):.2f}"
            else:
                ws.cell(row=row, column=4).value = "N/A"
                ws.cell(row=row, column=5).value = "N/A"
            
            # Determine winner
            if crewai_avg > 0 and langchain_avg > 0:
                winner = "CrewAI" if crewai_avg < langchain_avg else "LangChain"
                improvement = abs(crewai_avg - langchain_avg) / max(crewai_avg, langchain_avg) * 100
                ws.cell(row=row, column=6).value = f"{winner} (+{improvement:.1f}%)"
            else:
                ws.cell(row=row, column=6).value = "Insufficient data"
            
            row += 1
        
        # Framework summary
        row += 2
        ws[f'A{row}'] = "Framework Summary"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        summary_headers = ["Framework", "Documents Processed", "Average Total Time (ms)", "Success Rate"]
        for col, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        
        row += 1
        for framework, totals in framework_totals.items():
            ws.cell(row=row, column=1).value = framework.title()
            ws.cell(row=row, column=2).value = totals['document_count']
            
            avg_time = totals['total_time_ms'] / totals['document_count'] if totals['document_count'] > 0 else 0
            ws.cell(row=row, column=3).value = f"{avg_time:.2f}"
            ws.cell(row=row, column=4).value = "100%"  # Assuming success if we have trace data
            
            row += 1
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)

    def export_events_json(self, output_path: str) -> str:
        """Export all events to JSON file"""
        all_events = self.collector.get_events()
        events_data = [event.to_dict() for event in all_events]
        
        export_data = {
            'session_info': self.collector.get_session_summary(),
            'events': events_data,
            'export_timestamp': datetime.now(timezone.utc).isoformat()
        }
        
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w') as f:
            json.dump(export_data, f, indent=2, default=str)
        
        return str(output_path)
    
    def record_classification(self, implementation: ImplementationType, document_path: str,
                            document_type: str, confidence: float, reasoning: str = None) -> str:
        """Record a classification event with structured result"""
        classification_data = {
            'document_type': document_type,
            'confidence': confidence,
            'reasoning': reasoning or f"Classified as {document_type}"
        }
        
        return self.log_stage_event(
            implementation=implementation,
            document_path=document_path,
            stage="classification",
            event_type=EventType.CLASSIFICATION_END,
            success=True,
            metadata=classification_data
        )
    
    def record_extraction(self, implementation: ImplementationType, document_path: str,
                         extraction_result: Dict[str, Any]) -> str:
        """Record an extraction event with structured result"""
        return self.log_stage_event(
            implementation=implementation,
            document_path=document_path,
            stage="extraction",
            event_type=EventType.EXTRACTION_END,
            success=True,
            metadata=extraction_result
        )
    
    def record_accuracy_assessment(self, implementation: ImplementationType, document_path: str,
                                 accuracy_result: Dict[str, Any]) -> str:
        """Record an accuracy assessment event with structured result"""
        return self.log_stage_event(
            implementation=implementation,
            document_path=document_path,
            stage="accuracy_assessment",
            event_type=EventType.ACCURACY_ASSESSMENT_END,
            success=True,
            metadata=accuracy_result
        )
    
    def record_enhanced_result(self, implementation: ImplementationType, document_path: str,
                             enhanced_result: Dict[str, Any]) -> str:
        """Record an enhanced result with complete trace route information"""
        # Extract trace route information
        processing_pipeline = enhanced_result.get('processing_pipeline', [])
        complete_trace = enhanced_result.get('complete_trace_route', {})
        
        # Create event with enhanced metadata
        metadata = {
            'enhanced_result': enhanced_result,
            'processing_pipeline': processing_pipeline,
            'complete_trace_route': complete_trace,
            'step_details': enhanced_result.get('step_details', {}),
            'timing_details': enhanced_result.get('timing_details', {}),
            'classification_json': enhanced_result.get('classification_json'),
            'extraction_json': enhanced_result.get('extraction_json'), 
            'accuracy_json': enhanced_result.get('accuracy_json'),
            'final_decision_json': enhanced_result.get('final_decision_json')
        }
        
        return self.log_stage_event(
            implementation=implementation,
            document_path=document_path,
            stage="enhanced_processing_complete",
            event_type=EventType.PROCESSING_END,
            success=enhanced_result.get('status') == 'success',
            metadata=metadata
        )
    
    def export_enhanced_json_results(self, output_path: str) -> str:
        """Export enhanced JSON results with complete trace routes"""
        all_events = self.collector.get_events()
        
        # Group events by document and implementation
        results_by_document = {}
        
        for event in all_events:
            if event.event_type == EventType.PROCESSING_END and event.metadata.get('enhanced_result'):
                doc_key = f"{event.document_name}_{event.implementation.value}"
                results_by_document[doc_key] = {
                    'document_path': event.document_path,
                    'document_name': event.document_name,
                    'implementation': event.implementation.value,
                    'timestamp': event.timestamp.isoformat(),
                    'enhanced_result': event.metadata['enhanced_result'],
                    'processing_pipeline': event.metadata.get('processing_pipeline', []),
                    'complete_trace_route': event.metadata.get('complete_trace_route', {}),
                    'classification_json': event.metadata.get('classification_json'),
                    'extraction_json': event.metadata.get('extraction_json'),
                    'accuracy_json': event.metadata.get('accuracy_json'),
                    'final_decision_json': event.metadata.get('final_decision_json')
                }
        
        # Create comprehensive export data
        export_data = {
            'export_info': {
                'export_timestamp': datetime.now(timezone.utc).isoformat(),
                'total_documents_processed': len(set(r['document_name'] for r in results_by_document.values())),
                'total_processing_sessions': len(results_by_document),
                'frameworks_used': list(set(r['implementation'] for r in results_by_document.values()))
            },
            'session_summary': self.collector.get_session_summary(),
            'processing_results': results_by_document,
            'trace_route_analysis': self._analyze_trace_routes(results_by_document),
            'performance_comparison': self._compare_framework_performance(results_by_document)
        }
        
        # Save to file
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)
        
        return str(output_path)
    
    def _analyze_trace_routes(self, results_by_document: Dict[str, Any]) -> Dict[str, Any]:
        """Analyze trace routes across all processing sessions"""
        analysis = {
            'total_steps_across_all_sessions': 0,
            'average_processing_time_ms': 0,
            'step_performance_breakdown': {},
            'framework_comparison': {},
            'common_bottlenecks': []
        }
        
        all_pipelines = []
        for result in results_by_document.values():
            pipeline = result.get('processing_pipeline', [])
            if pipeline:
                all_pipelines.append({
                    'framework': result['implementation'],
                    'pipeline': pipeline,
                    'total_time_ms': sum(step.get('duration_ms', 0) for step in pipeline)
                })
        
        if all_pipelines:
            # Calculate averages
            total_time = sum(p['total_time_ms'] for p in all_pipelines)
            analysis['average_processing_time_ms'] = total_time / len(all_pipelines)
            
            # Step breakdown
            step_times = {}
            for pipeline_data in all_pipelines:
                for step in pipeline_data['pipeline']:
                    step_name = step.get('step', 'unknown')
                    duration = step.get('duration_ms', 0)
                    if step_name not in step_times:
                        step_times[step_name] = []
                    step_times[step_name].append(duration)
            
            for step_name, times in step_times.items():
                analysis['step_performance_breakdown'][step_name] = {
                    'average_duration_ms': sum(times) / len(times),
                    'min_duration_ms': min(times),
                    'max_duration_ms': max(times),
                    'occurrence_count': len(times)
                }
        
        return analysis
    
    def _compare_framework_performance(self, results_by_document: Dict[str, Any]) -> Dict[str, Any]:
        """Compare performance between frameworks"""
        framework_stats = {}
        
        for result in results_by_document.values():
            framework = result['implementation']
            if framework not in framework_stats:
                framework_stats[framework] = {
                    'total_documents': 0,
                    'total_processing_time_ms': 0,
                    'successful_processes': 0,
                    'average_accuracy': 0,
                    'accuracy_scores': []
                }
            
            stats = framework_stats[framework]
            stats['total_documents'] += 1
            
            # Get timing information
            pipeline = result.get('processing_pipeline', [])
            total_time = sum(step.get('duration_ms', 0) for step in pipeline)
            stats['total_processing_time_ms'] += total_time
            
            # Get accuracy information
            enhanced_result = result.get('enhanced_result', {})
            accuracy = enhanced_result.get('overall_accuracy', 0)
            if accuracy > 0:
                stats['accuracy_scores'].append(accuracy)
            
            # Check success
            if enhanced_result.get('status') == 'success':
                stats['successful_processes'] += 1
        
        # Calculate derived metrics
        for framework, stats in framework_stats.items():
            if stats['total_documents'] > 0:
                stats['average_processing_time_ms'] = stats['total_processing_time_ms'] / stats['total_documents']
                stats['success_rate'] = stats['successful_processes'] / stats['total_documents']
                
                if stats['accuracy_scores']:
                    stats['average_accuracy'] = sum(stats['accuracy_scores']) / len(stats['accuracy_scores'])
        
        return framework_stats

    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content with special handling for JSON columns"""
        from openpyxl.utils import column_index_from_string
        
        for column in ws.columns:
            max_length = 0
            column_letter = None
            
            # Find the first regular cell to get column letter
            for cell in column:
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                    break
            
            if column_letter is None:
                continue
                
            # Calculate max width for this column
            for cell in column:
                try:
                    if hasattr(cell, 'value') and cell.value is not None:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            
            # Enhanced width for JSON columns - provide extra space for pretty-printed JSON
            column_index = column_index_from_string(column_letter)
            if column_index >= 14 and column_index <= 17:  # JSON columns (Classification, Extraction, Accuracy, Final Decision)
                adjusted_width = min(max(max_length * 0.15, 60), 120)  # Larger minimum and maximum for JSON readability
            else:
                adjusted_width = min(max_length + 2, 50)  # Standard width for other columns
            
            ws.column_dimensions[column_letter].width = adjusted_width
